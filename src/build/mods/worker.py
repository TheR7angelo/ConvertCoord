import csv
import locale
import os
import datetime
import re
import shutil
import sqlite3
import sys
import csv
from collections import defaultdict
from pprint import pprint

import openpyxl
import psycopg2
import unicodedata
from PySide6.QtCore import QObject, Signal, QRunnable
from PySide6.QtWidgets import QTableWidgetItem
from openpyxl import load_workbook
from pyproj import Transformer

bdd = r"T:\- 4 Suivi Appuis\18-Partage\BARRENTO ANTUNES Raphael\6_Programme python\Instalateur\Export apcom\bdd.db"

class WorkerSignals(QObject):
    start = Signal(bool)
    progress = Signal(int)
    erreur = Signal(str)
    data_dict = Signal(dict)
    data_liste = Signal(list)
    max_value = Signal(int)
    finish = Signal(bool)


class Worker_export(QRunnable):
    signals = WorkerSignals()
    table_header = []

    def __init__(self, files, fonction, marcher, table):
        super(Worker_export, self).__init__()

        self.fonction = fonction
        self.files = files
        self.marcher = marcher
        self.table = table
        self.psql = shutil.which("psql")

    def run(self):  # sourcery no-metrics

        def presence(id_unique):
            id_unique = id_unique.replace("'", "''")

            with sqlite3.connect(bdd) as conn:
                cursor = conn.cursor()

                if cursor.execute(f"SELECT * FROM t_apcom WHERE ap_id='{id_unique}';").fetchone():
                    return "OUI"
                else:
                    return "NON"

        def C6_Type(geospot: str):
            if geospot[0] in ['B', '1']:
                return 'BOIS'
            elif geospot[0] == 'M':
                return 'METAL'
            elif geospot[0] == 'F':
                return 'COMPOSITE'
            elif geospot[0].lower() in ['c', 'd', "e"]:
                return 'BETON'
            elif geospot[0].lower() == 'p':
                return 'POTELET'
            else:
                return "NON DEFINI"

        def info_bt(workbook):  # sourcery no-metrics
            liste_bt = []
            worksheet = workbook.worksheets[0]
            for rows in worksheet.iter_rows(min_row=4, min_col=1, max_col=1):
                for cell in rows:
                    if cell.value is not None:
                        coord = str(worksheet[f'K{str(cell.row)}'].value), str(
                            worksheet[f'L{str(cell.row)}'].value
                        )
                        nombre_cable = str(worksheet[f'AO{cell.row}'].value).replace('None', '')
                        ML = str(worksheet[f'AU{cell.row}'].value)

                        if not nombre_cable:
                            # Utilisation = 'D_3'
                            Utilisation = 'D_2'
                            ML = '20'
                        else:
                            Utilisation = 'D_1' if '15' in nombre_cable else 'D_2'
                        Resultat = str(worksheet[f'AV{cell.row}'].value)
                        # Malt = str(worksheet[f'S{str(cell.row)}'].value)
                        # BoitierTelecom = str(worksheet[f'AD{str(cell.row)}'].value)

                        if Resultat == 'OK':
                            Resultat = 'SATISFAISANT'
                            # Travaux = 'AUCUN'
                            Travaux = ""
                            Bandeu_vert = "OUI"
                            # if Utilisation == 'D_3':
                            #     Observation = f'ENEC {str(inc + 1)} ENEPB'
                            #     inc += 1
                            # else:
                            #     Observation = 'Bandeau vert OK'
                        else:
                            Resultat = 'INSUFFISANT'
                            Travaux = 'IMPLANTATION'
                            Bandeu_vert = "NON"
                            # if Utilisation == 'D_3':
                            #     Observation = f'ENEC {str(inc)}'
                            #     inc += 1
                            # else:
                            #     Observation = ''
                        tmp = {
                            'Propriétaire': 'ENEDIS',
                            'Partenaire': 'INEO',
                            'Armoire FI': "",
                            'Armoire PA': "",
                            'Nom Appui': str(worksheet[f'A{cell.row}'].value).strip(),
                            'Adresse (rue)': str(worksheet['Q1'].value).strip(),
                            'Code Insee': str(worksheet['I1'].value).strip(),
                            'EP': str(worksheet[f'Z{cell.row}'].value).replace('None', 'NON').strip(),
                            'Boîtier FTTH': str(worksheet[f'AR{cell.row}'].value).strip(),
                            'Latitude': str(coord[0]).strip(),
                            'Longitude': str(coord[1]).strip(),
                            'Nombre / nature câble': str(nombre_cable).strip(),
                            'Résultat étude': str(Resultat).strip(),
                            'Métré linéaire': str(ML).strip(),
                            'Travaux': str(Travaux).strip(),
                            'Utilisation': str(Utilisation).strip(),
                            'Présence terre': 'NON'.strip(),
                            'Tension électrique': str(worksheet[f'C{cell.row}'].value).strip(),
                            'Boitier telecom': 'NON'.strip(),
                            'Référence opérateur': "",
                            # 'Observations': Observation,
                            'Hauteur totale': str(worksheet[f'E{cell.row}'].value).strip(),
                            'Classe': str(worksheet[f'H{cell.row}'].value).strip(),
                            'Effort nominal': str(worksheet[f'I{cell.row}'].value).strip(),
                            'Bandeau vert': str(Bandeu_vert).strip(),
                            'Artère': "",
                            'PM': "",
                            "Effort": str(worksheet[f'AW{cell.row}'].value).strip()
                        }
                        liste_bt.append(tmp.copy())
            return liste_bt

        def info_ft(workbook):  # sourcery no-metrics
            liste_ft = []

            worksheet = workbook.worksheets[len(workbook.worksheets) - 1]
            for rows in worksheet.iter_rows(min_row=9, min_col=19, max_col=19):
                for cell in rows:
                    print(cell)
                    if cell.value is not None and worksheet[f"A{cell.row}"].value is not None:
                        type_appuis = C6_Type(str(worksheet[f"B{cell.row}"].value))
                        if type_appuis not in ["BETON", "POTELET"]:
                            type_appuis_2 = str(worksheet[f"B{cell.row}"].value)
                            Proprietaire = 'ORANGE'
                            Partenaire = 'INEO'
                            NomAppui = str(worksheet[f"A{cell.row}"].value)
                            while len(NomAppui) < 6:
                                NomAppui = f"0{NomAppui}"
                            Adresse = worksheet[f"C{cell.row}"].value
                            INSEE = str(worksheet['G3'].value)
                            EP = 'NON'
                            PB = 'OUI' if worksheet[f"AJ{cell.row}"].value in ("PB", "OUI") else 'NON'
                            remplacement = "" if worksheet[f"AG{cell.row}"].value is None else worksheet[f"AG{cell.row}"].value
                            x, y = (
                                worksheet[f'D{str(cell.row)}'].value,
                                worksheet[f'E{str(cell.row)}'].value,
                            )
                            coord = correcteur("capft", x, y)

                            if worksheet[f"AD{cell.row}"].value == '' or worksheet[f"AD{cell.row}"].value is None :
                                ResultatEtude = "SATISFAISANT"
                            elif worksheet[f"AC{cell.row}"].value is None or int(worksheet[f"AC{cell.row}"].value) > 0:
                                ResultatEtude = "SATISFAISANT"
                            else:
                                ResultatEtude = "INSUFFISANT"

                            TypeTravaux = worksheet[f"AF{cell.row}"].value
                            Utilisable = False if worksheet[f"AD{cell.row}"].fill is None else worksheet[f"AD{cell.row}"].fill.start_color.index not in (50, 53, "FF92D050", "FFFAAA46")

                            if TypeTravaux is None or TypeTravaux == '':
                                # TypeTravaux = 'AUCUN'
                                TypeTravaux = ""
                            elif Utilisable:
                                TypeTravaux = "IMPLANTATION"
                                ResultatEtude = "INSUFFISANT"
                            else:
                                TypeTravaux = 'REMPLACEMENT'
                                ResultatEtude = "INSUFFISANT"

                            if (worksheet[f"AK{cell.row}"].value is not None or worksheet[
                                f"AK{cell.row}"].value != '') or worksheet[f"AK{cell.row}"].value == 'NON' or worksheet[
                                f"AK{cell.row}"].value == '0':
                                BoitierTelecom = "NON"
                            else:
                                BoitierTelecom = "OUI"
                            affaire_operateur = ""

                            start = True
                            nombre_cable = []
                            ML = 0
                            while (worksheet[f"A{cell.row}"].value is None or start) and worksheet[
                                f"S{cell.row}"].value is not None:
                                start = False
                                if "L10" in worksheet[f"S{cell.row}"].value:
                                    nombre_cable.append(worksheet[f"S{cell.row}"].value)
                                    ML += worksheet[f"T{cell.row}"].value
                                cell.row += 1
                            if not nombre_cable:
                                # Utilisation = "D_3"
                                Utilisation = "D_2"
                            else:
                                Utilisation = "D_1" if "15" in nombre_cable else "D_2"
                            tmp = {
                                'Propriétaire': Proprietaire.strip(),
                                'Partenaire': Partenaire.strip(),
                                'Armoire FI': "",
                                'Armoire PA': "",
                                'Nom Appui': NomAppui.strip(),
                                'Adresse (rue)': Adresse.strip(),
                                'Code Insee': INSEE.strip(),
                                'EP': EP.strip(),
                                'Boîtier FTTH': PB.strip(),
                                'Latitude': f"{coord[0]}".strip(),
                                'Longitude': f"{coord[1]}".strip(),
                                'Nombre / nature câble': "||".join(nombre_cable).strip(),
                                'Résultat étude': ResultatEtude.strip(),
                                'Métré linéaire': str(round(ML / 2, 2)).replace('.', ',').strip(),
                                'Travaux': TypeTravaux.strip(),
                                'Utilisation': Utilisation.strip(),
                                'Boitier telecom': BoitierTelecom.strip(),
                                'Référence opérateur': affaire_operateur.strip(),
                                'Référence GESPOT': type_appuis_2.strip(),
                                "Poteaux replacement": remplacement.strip(),
                                'Artère': "",
                                'PM': ""
                            }
                            liste_ft.append(tmp.copy())
                            tmp.clear()
            return liste_ft

        def correcteur(logiciel, coord_x, coord_y):
            """
            :param logiciel: capft || comac
            :return: coord_x coord_y
            """

            def wgs_dms(coord_x, coord_y):
                wsg = []
                coord = [float(coord_x), float(coord_y)]
                for index, value in enumerate(coord):

                    if value > 0:
                        sens = '+'
                    else:
                        sens = '-'
                        value = -value

                    lonDegrees = value
                    lonSeconds = lonDegrees * 60 * 60

                    seconds = float(lonSeconds % 60)
                    minutes = int((lonSeconds / 60) % 60)
                    degrees = int((lonSeconds / 60) / 60)

                    if sens == '-':
                        sens = 'W' if index == 0 else 'S'
                    else:
                        sens = 'E' if index == 0 else 'N'
                    wsg.append(f'{degrees}°{minutes}' + "'" + str("%.4f" % seconds) + '"' + sens)

                return wsg[1], wsg[0]

            def lambert_dms(coord_x, coord_y):
                transformer = Transformer.from_crs("epsg:2154", "epsg:4326", always_xy=True)

                try:
                    coord_x, coord_y = transformer.transform(coord_x, coord_y)
                    coord_x, coord_y = wgs_dms(coord_x, coord_y)
                except Exception:
                    pass

                return coord_x, coord_y

            if logiciel == "capft":

                try:
                    coord_x = float(coord_x.replace(",", "."))
                except Exception:
                    pass

                try:
                    coord_y = float(coord_y.replace(",", "."))
                except Exception:
                    pass

                if not isinstance(coord_x, str):
                    if -180 <= int(coord_x) <= 180 and -90 < int(coord_y) < 90:
                        coord = wgs_dms(coord_x, coord_y)
                    else:
                        coord = lambert_dms(coord_x, coord_y)
                    return coord[0], coord[1]
                else:
                    try:
                        coord_x, coord_y = correcteur(logiciel, float(coord_x), float(coord_y))
                    except Exception:
                        pass
                    return coord_x, coord_y
            else:
                return coord_x, coord_y

        def converter(proj, coord_x, coord_y):  # sourcery no-metrics
            def sexa_to_dms(coord):
                coord = re.split('°|\'|"', coord)
                coord = [item.strip() for item in coord]

                coord = [float(item) if item.replace('.', '', 1).isnumeric() else item for item in coord]

                if coord[3].lower() in ["n", "e"]:
                    return coord[0] + (coord[1] / 60) + (coord[2] / 3600)
                else:
                    return -(coord[0] + (coord[1] / 60) + (coord[2] / 3600))

            try:
                coord_x = float(coord_x.replace(",", "."))
            except Exception:
                pass

            try:
                coord_y = float(coord_y.replace(",", "."))
            except Exception:
                pass

            if proj == "comac":
                transformer = Transformer.from_crs("epsg:2154", "epsg:4326", always_xy=True)
                coord_x, coord_y = transformer.transform(coord_x, coord_y)
            elif proj == "capft":
                if isinstance(coord_x, str):
                    if "n" in coord_x.lower() or "s" in coord_x.lower() or "n" in coord_y.lower() or "s" in coord_y.lower():
                        coord_x, coord_y = sexa_to_dms(coord_x), sexa_to_dms(coord_y)
                        transformer = Transformer.from_crs("epsg:4326", "epsg:4326", always_xy=True)
                    elif -180 <= int(coord_x) <= 180 and -90 <= int(coord_y) < 90:
                        transformer = Transformer.from_crs("epsg:4326", "epsg:4326", always_xy=True)
                    else:
                        transformer = Transformer.from_crs("epsg:2154", "epsg:4326", always_xy=True)
                    try:
                        coord_x, coord_y = transformer.transform(coord_x, coord_y)
                    except Exception:
                        pass
                else:
                    if -180 <= int(coord_x) <= 180 and -90 <= int(coord_y) < 90:
                        transformer = Transformer.from_crs("epsg:4326", "epsg:4326", always_xy=True)
                    else:
                        transformer = Transformer.from_crs("epsg:2154", "epsg:4326", always_xy=True)
                    try:
                        coord_x, coord_y = transformer.transform(float(coord_x), float(coord_y))
                    except Exception:
                        pass

            return coord_x, coord_y

        def find_info(coord_x, coord_y, insee):

            if self.marcher == "amel":
                transformer = Transformer.from_crs("epsg:4326", "epsg:27582", always_xy=True)
                database = "AMEL_47"
            elif self.marcher == "rip":
                insee = str(insee)[:2]
                database = f"RIP_FTTH_{insee}"
                transformer = Transformer.from_crs("epsg:4326", "epsg:2154", always_xy=True)
            else:
                transformer = Transformer.from_crs("epsg:4326", "epsg:27582", always_xy=True)
                database = "CEM_ORANGE"
            coord_x, coord_y = transformer.transform(coord_x, coord_y)

            try:
                with psycopg2.connect(database=database, user="postgres", password="INEO_Infracom_33",
                                      host="bordeaux04", port="5432") as conn:
                    cursor = conn.cursor()

                    if self.marcher in ["amel", "cem"]:
                        cursor.execute(f"""
                                WITH pmz AS (
                                    SELECT zo.dept, zo.id_metier_ AS zone_pm
                                    FROM geofibre_cem.zone_eligibilite_cem AS zo
                                    WHERE ST_Intersects(st_geometryfromtext('multipoint({coord_x} {coord_y})', 27582) , zo.geom)
                                    AND zo.type_pf IN ('PMZ')
                                    )
                                , pa AS (
                                    SELECT zo.dept, zo.id_metier_ AS zone_pa, zo.code_com as code_com
                                    FROM geofibre_cem.zone_eligibilite_cem AS zo
                                    WHERE ST_Intersects(st_geometryfromtext('multipoint({coord_x} {coord_y})', 27582) , zo.geom)
                                    AND zo.type_pf IN ('PA')
                                    )
                                ,armoire AS (
                                    SELECT DISTINCT zo.pa, zo.pm_regleme, zo.pa_regleme
                                    FROM public.v_contour_aerien_cem AS zo
                                )
                                SELECT pmz.dept, pmz.zone_pm, pa.zone_pa, armoire.pm_regleme, armoire.pa_regleme
                                FROM pmz
                                LEFT JOIN pa
                                ON pa.dept = pmz.dept
                                LEFT JOIN armoire
                                ON armoire.pa=pa.zone_pa
                                """)
                    elif self.marcher == "rip":
                        cursor.execute(f"""
                                        WITH pmz AS (
                                            SELECT zo.dept, zo.id_metier_ AS zone_pm
                                            FROM telecom.zone_eligibilite AS zo
                                            WHERE ST_Intersects(st_geometryfromtext('multipoint({coord_x} {coord_y})', 2154), zo.the_geom)
                                            AND zo.type_pf IN ('PMZ')
                                            )
                                        ,pa AS (
                                            SELECT zo.dept, zo.id_metier_ AS zone_pa, zo.code_com as code_com
                                            FROM telecom.zone_eligibilite AS zo
                                            WHERE ST_Intersects(st_geometryfromtext('multipoint({coord_x} {coord_y})', 2154), zo.the_geom)
                                            AND zo.type_pf IN ('PA')
                                        )
                                        ,armoire AS (
                                            SELECT DISTINCT zo.pa, zo.pm_regleme, zo.pa_regleme
                                            FROM suivi.v_contour_aerien AS zo
                                        )
                                        SELECT pmz.dept, pmz.zone_pm, pa.zone_pa, armoire.pm_regleme, armoire.pa_regleme
                                        FROM pmz
                                        LEFT JOIN pa
                                        ON pa.dept = pmz.dept
                                        LEFT JOIN armoire
                                        ON armoire.pa=pa.zone_pa
                                        """)

                    row = cursor.fetchone()
                    col = [i[0] for i in cursor.description]

                    value = dict(zip(col, row))
            except Exception:
                value = None

            return value

        def type_file(workbook, path):

            path = path.split("\\")
            path = [item.lower() for item in path[-4:]]
            if "old" in path:
                return "Old", "False"
            sheet = [worksheet.title for worksheet in workbook.worksheets]

            if len(sheet) == 1 and sheet[0] == "Etude Réseau Télécom":
                version = "False"
                for sub_path in path:
                    if "solution" in sub_path:
                        version = "solution"
                        break
                    elif "initial" in sub_path:
                        version = "initial"
                        break
                return "comac", version
            elif len(sheet) == 1 or len(sheet) != 4:
                return "Autre", "False"
            else:
                return (
                    ("capft", "True")
                    if sheet[0] == "Saisies terrain"
                       and sheet[1] == "Bases"
                       and sheet[2] == "Photos"
                       and sheet[3] == "Export 1"
                    else ("Autre", "False")
                )

        self.signals.start.emit(True)

        try:

            for i in range(self.table.columnCount()):
                self.table_header.append(self.table.horizontalHeaderItem(i).text())

            locale.setlocale(locale.LC_TIME, "fr_FR")

            row = 0

            self.signals.max_value.emit(len(self.files))

            for index, file in enumerate(self.files):
                dossier, fichier = os.path.split(file)
                if not "~$" in file:
                    print(file)
                    workbook = load_workbook(filename=file, read_only=True, data_only=True, keep_vba=False)

                    logiciel, version = type_file(workbook, file)
                    print(logiciel)

                    if logiciel == "comac":
                        worksheet = workbook.worksheets[0]
                        # coord_x, coord_y = worksheet.cell(row=4, column=11).value, worksheet.cell(row=4,
                        #                                                                           column=12).value
                        # coord_x, coord_y = converter(proj="comac", coord_x=coord_x, coord_y=coord_y)
                        insee = worksheet.cell(row=1, column=9).value
                        list_appuis = info_bt(workbook)
                    elif logiciel == "capft":
                        worksheet = workbook.worksheets[3]
                        # coord_x, coord_y = worksheet.cell(row=9, column=4).value, worksheet.cell(row=9, column=5).value
                        # coord_x, coord_y = converter(proj="capft", coord_x=coord_x, coord_y=coord_y)
                        insee = worksheet.cell(row=3, column=7).value
                        list_appuis = info_ft(workbook)
                    else:
                        insee = None
                        list_appuis = []
                    workbook.close()

                    if len(list_appuis) > 0:
                        info = None
                        for appuis in list_appuis:

                            coord_x, coord_y = appuis["Longitude"], appuis["Latitude"]
                            if isinstance(coord_x, str):
                                if "w" in coord_y.lower() or "e" in coord_y.lower():
                                    coord_x, coord_y = coord_y, coord_x

                            if coord_x > coord_y:
                                coord_x, coord_y = coord_y, coord_x
                            coord_x, coord_y = converter(proj=logiciel, coord_x=coord_x, coord_y=coord_y)

                            print(coord_x, coord_y)

                            info = find_info(coord_x=coord_x, coord_y=coord_y, insee=insee)
                            if info is not None:
                                if info["zone_pa"] is not None:
                                    break

                        if info is None:
                            if self.psql:
                                info = defaultdict(lambda: "Erreur coordonnée")
                            else:
                                info = defaultdict(lambda: "PostgreSQL n'est pas installer")

                        for appuis in list_appuis:
                            self.table.insertRow(row)
                            for key in appuis:
                                col = self.table_header.index(key)

                                if key == "Armoire FI":
                                    self.table.setItem(row, col, QTableWidgetItem(info["pm_regleme"]))
                                elif key == "Armoire PA":
                                    self.table.setItem(row, col, QTableWidgetItem(info["pa_regleme"]))
                                elif key == "Référence opérateur":
                                    self.table.setItem(row, col, QTableWidgetItem(fichier.split(".")[0]))
                                elif key == "Artère":
                                    self.table.setItem(row, col, QTableWidgetItem(file))
                                elif key == "PM":
                                    self.table.setItem(row, col, QTableWidgetItem(info["zone_pm"]))
                                else:
                                    self.table.setItem(row, col, QTableWidgetItem(appuis[key]))
                            col = self.table_header.index("Doublon APCOM")
                            redepot = presence(id_unique=f"{appuis['Nom Appui']}/{appuis['Code Insee']}/{appuis['Adresse (rue)']}")
                            self.table.setItem(row, col, QTableWidgetItem(redepot))
                            row += 1
                        list_appuis.clear()

                self.signals.progress.emit(index + 1)

        except Exception as error:
            error = f"Ligne {sys.exc_info()[-1].tb_lineno} {error}"
            print(error)

            date = datetime.datetime.now().strftime('%d-%m-%Y %H-%M-%S')
            file = os.path.basename(__file__).replace('.py', '')
            directory = fr"src/log/crash/{self.fonction}/"

            os.makedirs(directory, exist_ok=True)
            with open(f'{directory}{file}_{date}.txt', 'w') as crash_report:
                crash_report.write(str(error))
            os.startfile(os.path.abspath(f'{directory}{file}_{date}.txt'))

        self.signals.start.emit(False)


class Worker_export_file(QRunnable):
    signals = WorkerSignals()

    def __init__(self, fonction, table, path, marcher):
        super(Worker_export_file, self).__init__()

        self.fonction = fonction
        self.table = table
        self.path = path
        self.marcher = marcher

    def run(self):  # sourcery no-metrics

        self.signals.start.emit(True)

        ext = os.path.split(self.path)[1].split(".")[1]

        try:

            workbook = openpyxl.Workbook()
            workbook.properties.creator = os.environ.get('USERNAME')
            worksheet = workbook.worksheets[0]
            worksheet.title = 'Export'

            max = 1
            row_file = 1

            for col_file, i in enumerate(range(self.table.columnCount() - max), start=1):
                item = self.table.horizontalHeaderItem(i).text()

                cell = worksheet.cell(row=row_file, column=col_file)
                cell.value = item
            row_file += 1
            col_file = 1

            for row_table in range(self.table.rowCount()):
                for col_table in range(self.table.columnCount() - max):
                    try:
                        item = self.table.item(row_table, col_table).text()
                    except AttributeError:
                        item = ""

                    cell = worksheet.cell(row=row_file, column=col_file)
                    cell.value = item
                    col_file += 1

                row_file += 1
                col_file = 1



                progress = int(100 * (row_table + 1) / self.table.rowCount())
                if progress == 100:
                    progress = 99
                self.signals.progress.emit(progress)

            if ext == "xlsx":
                workbook.save(self.path)
            workbook.close()

            self.signals.progress.emit(100)

            os.startfile(self.path)

        except Exception as error:
            error = f"Ligne {sys.exc_info()[-1].tb_lineno} {error}"
            print(error)

            date = datetime.datetime.now().strftime('%d-%m-%Y %H-%M-%S')
            file = os.path.basename(__file__).replace('.py', '')
            directory = fr"src/log/crash/{self.fonction}/"

            os.makedirs(directory, exist_ok=True)
            with open(f'{directory}{file}_{date}.txt', 'w') as crash_report:
                crash_report.write(str(error))

        self.signals.start.emit(False)


class Worker_export_bdd(QRunnable):
    signals = WorkerSignals()

    def __init__(self, fonction, table, path, marcher):
        super(Worker_export_bdd, self).__init__()

        self.fonction = fonction
        self.table = table
        self.path = path
        self.marcher = marcher

    def run(self):

        self.signals.start.emit(True)

        try:
            ext = os.path.split(self.path)[1].split(".")[1]
            header = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]
            date = datetime.datetime.now().strftime("%d/%m/%Y")
            workbook = openpyxl.Workbook()
            workbook.properties.creator = os.environ.get('USERNAME')
            worksheet = workbook.worksheets[0]
            worksheet.title = 'Export'

            if self.fonction == "minimum":
                header_workbook = [
                    'Propriétaire', 'Partenaire', 'Armoire FI', 'Armoire PA', 'Nom Appui', 'Adresse (rue)',
                    'Code Insee', 'EP', 'Boîtier FTTH', 'Latitude', 'Longitude', 'Nombre / nature câble',
                    'Résultat étude', 'Métré linéaire', 'Travaux', 'Utilisation', 'Présence terre',
                    'Tension électrique', 'Boitier telecom', 'Référence opérateur', 'Observations', 'Hauteur totale',
                    'Classe', 'Effort nominal', 'Référence GESPOT', 'Bandeau vert'
                ]
            else:
                header_workbook = [self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())]

            for index, item in enumerate(header_workbook):
                worksheet.cell(row=1, column=index+1).value = item

            with sqlite3.connect(bdd) as conn:
                cursor = conn.cursor()
                for row_table in range(self.table.rowCount()):
                    data = []
                    for col_table in range(self.table.columnCount()):
                        try:
                            item = self.table.item(row_table, col_table).text()
                        except AttributeError:
                            item = None
                        if item == "":
                            item = None
                        data.append(item)

                    data = dict(zip(header, data))

                    requete = bdd_req(data=data, date=date)
                    tmp_header = ", ".join(list(requete.keys()))
                    tmp_value = list(requete.values())

                    tmp_value = ", ".join(tmp_value)

                    num_appui = requete["ap_id"]
                    # print(f"SELECT * FROM t_apcom WHERE ap_id={num_appui}")
                    cursor.execute(f"SELECT * FROM t_apcom WHERE ap_id={num_appui}")
                    row = cursor.fetchone()
                    if row is None:
                        cursor.execute(f"INSERT INTO t_apcom({tmp_header}) VALUES({tmp_value});")
                        print(f'insert {requete["ap_id"]}')
                    else:
                        del requete["ap_id"]

                        update = ", ".join([f"{key}={value}" for key, value in requete.items()])
                        cursor.execute(f"UPDATE t_apcom SET {update} WHERE ap_id={num_appui};")
                        print(f"update {num_appui}")

                    for index, key in enumerate(header_workbook):
                        try:
                            worksheet.cell(row=row_table+2, column=index+1).value = data[key].replace("''", "'")
                        except Exception:
                            worksheet.cell(row=row_table + 2, column=index + 1).value = data[key]

                progress = int(100 * (row_table + 1) / self.table.rowCount())
                if progress == 100:
                    progress = 99
                self.signals.progress.emit(progress)

            if ext == "xlsx":
                workbook.save(self.path)
            elif ext == "csv":
                col = csv.writer(open(self.path, 'w', newline=""), quoting=csv.QUOTE_NONNUMERIC, delimiter=';')
                for r in worksheet.rows:
                    col.writerow([cell.value for cell in r])
                workbook.close()
            self.signals.progress.emit(100)

        except Exception as error:
            error = f"Ligne {sys.exc_info()[-1].tb_lineno} {error}"
            print(error)

            date = datetime.datetime.now().strftime('%d-%m-%Y %H-%M-%S')
            file = os.path.basename(__file__).replace('.py', '')
            directory = fr"log/crash/{self.fonction}/"

            os.makedirs(directory, exist_ok=True)
            with open(f'{directory}{file}_{date}.txt', 'w') as crash_report:
                crash_report.write(str(error))
            os.startfile(os.path.abspath(f'{directory}{file}_{date}.txt'))

        self.signals.start.emit(False)


class Worker_xlsx_bdd(QRunnable):
    signals = WorkerSignals()

    def __init__(self, files, table, fonction):
        super(Worker_xlsx_bdd, self).__init__()
        self.files = files
        self.table = table
        self.fonction = fonction

    def run(self):

        def strip_accents(s):
            return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')

        self.signals.start.emit(True)

        column_header = [self.table.horizontalHeaderItem(col).text() for col in range(self.table.columnCount())]

        try:
            for file in self.files:
                workbook = load_workbook(filename=file, read_only=True, keep_vba=False)
                worksheet = workbook.worksheets[0]

                header = []
                for ligne, row in enumerate(worksheet.iter_rows(min_row=1)):
                    if ligne == 0:
                        header = [cell.value for cell in row]
                    else:
                        value = [cell.value for cell in row]
                        data = dict(zip(header, value))

                        self.table.insertRow(ligne-1)
                        for key, value in data.items():
                            try:
                                value = value.replace("’", "'").replace("", " ")
                            except Exception:
                                pass
                            if key not in ["MARCHE", "DATE"] and value is not None:
                                index = column_header.index(key)
                                if index in (0, 1):
                                    value = strip_accents(value).upper()
                                self.table.setItem(ligne-1, index, QTableWidgetItem(str(value)))

        except Exception as error:
            error = f"Ligne {sys.exc_info()[-1].tb_lineno} {error}"
            print(error)

            date = datetime.datetime.now().strftime('%d-%m-%Y %H-%M-%S')
            file = os.path.basename(__file__).replace('.py', '')
            directory = fr"log/crash/{self.fonction}/"

            os.makedirs(directory, exist_ok=True)
            with open(f'{directory}{file}_{date}.txt', 'w') as crash_report:
                crash_report.write(str(error))
            os.startfile(os.path.abspath(f'{directory}{file}_{date}.txt'))

        self.signals.start.emit(False)


class Worker_csv_bdd(QRunnable):
    signals = WorkerSignals()

    def __init__(self, file, table, fonction, marcher):
        super(Worker_csv_bdd, self).__init__()
        self.file = file
        self.table = table
        self.fonction = fonction
        self.marcher = marcher

    def run(self):
        self.signals.start.emit(True)

        try:
            date = datetime.datetime.now().strftime("%d/%m/%Y")

            with open(self.file, 'r', newline="") as csvfile:
                reader = csv.reader(csvfile, delimiter=';')
                rows = list(reader)
                header = rows[0]
                rows = rows[1:]

                rows = [dict(zip(header, row)) for row in rows]

                with sqlite3.connect(bdd) as conn:
                    cursor = conn.cursor()
                    for index, row in enumerate(rows):

                        requete = self.bdd_req(data=row, date=date)
                        tmp_header = ", ".join(list(requete.keys()))
                        tmp_value = list(requete.values())

                        tmp_value = ", ".join(tmp_value)

                        num_appui = requete["ap_id"]
                        cursor.execute(f"SELECT * FROM t_apcom WHERE ap_id={num_appui}")
                        row = cursor.fetchone()
                        if row is None:
                            cursor.execute(f"INSERT INTO t_apcom({tmp_header}) VALUES({tmp_value});")
                            print(f'insert {requete["ap_id"]}')
                        else:
                            del requete["ap_id"]

                            update = ", ".join([f"{key}={value}" for key, value in requete.items()])
                            cursor.execute(f"UPDATE t_apcom SET {update} WHERE ap_id={num_appui};")
                            print(f"update {num_appui}")

                        progress = int(100 * (index + 1) / len(rows))
                        if progress == 100:
                            progress = 99
                        self.signals.progress.emit(progress)

        except Exception as error:
            error = f"Ligne {sys.exc_info()[-1].tb_lineno} {error}"
            print(error)

            date = datetime.datetime.now().strftime('%d-%m-%Y %H-%M-%S')
            file = os.path.basename(__file__).replace('.py', '')
            directory = fr"log/crash/{self.fonction}/"

            os.makedirs(directory, exist_ok=True)
            with open(f'{directory}{file}_{date}.txt', 'w') as crash_report:
                crash_report.write(str(error))
            os.startfile(os.path.abspath(f'{directory}{file}_{date}.txt'))

        self.signals.start.emit(False)
        self.signals.finish.emit(True)

    def bdd_req(self, data, date):  # sourcery no-metrics
        def req(data, date):
            for key in data:
                try:
                    data[key] = data[key].replace("'", "''")
                except Exception:
                    pass

            bdd = {"ap_id": f"{data['Nom Appui']}/{data['Code Insee']}/{data['Adresse (rue)']}",
                   "ap_pro": f"(SELECT pr_id FROM t_pro WHERE pr_nom='{data['Propriétaire']}')",
                   "ap_part": f"(SELECT pt_id FROM t_part WHERE pt_nom='{data['Partenaire']}')",
                   "ap_fi": data["Armoire FI"],
                   "ap_pa": data["Armoire PA"],
                   "ap_nom": data['Nom Appui'],
                   "ap_adresse": data["Adresse (rue)"],
                   "ap_insee": int(data["Code Insee"]),
                   "ap_ep": f"(SELECT rp_id FROM t_yes_no WHERE rp_nom='{data['EP'].upper()}')",
                   "ap_ftth": f"(SELECT rp_id FROM t_yes_no WHERE rp_nom='{data['Boîtier FTTH'].upper()}')",
                   "ap_lat": data["Latitude"],
                   "ap_lon": data["Longitude"],
                   "ap_cable": data["Nombre / nature câble"],
                   "ap_result": f"(SELECT rs_id FROM t_result WHERE rs_nom='{data['Résultat étude']}')",
                   "ap_lineaire": float(data["Métré linéaire"].replace(",", ".")),
                   "ap_tvx": data['Travaux'],
                   "ap_n_app": None,
                   "ap_utilisation": f"(SELECT ut_id FROM t_utili WHERE ut_type='{data['Utilisation']}')",
                   "ap_terre": data["Présence terre"],
                   "ap_tension": data["Tension électrique"],
                   "ap_telecom": f"(SELECT rp_id FROM t_yes_no WHERE rp_nom='{data['Boitier telecom'].upper()}')",
                   "ap_refop": data["Référence opérateur"].replace("’", "''"),
                   "ap_obs": data["Observations"],
                   "ap_hauteur": data["Hauteur totale"],
                   "ap_class": data["Classe"],
                   "ap_effortnomi": data["Effort nominal"],
                   "ap_gespot": data["Référence GESPOT"],
                   "ap_bv": data["Bandeau vert"],
                   "ap_artere": None,
                   "ap_pm": None,
                   "ap_pourcent": None,
                   "ap_marcher": f"(SELECT mc_id FROM t_marche WHERE mc_nom='{self.marcher.upper()}_{data['Code Insee'][:2]}')",
                   "ap_date": date}

            try:
                bdd["ap_n_app"] = data["Poteaux replacement"]
            except Exception:
                pass
            try:
                bdd["ap_artere"] = data["Artère"]
            except Exception:
                pass
            try:
                bdd["ap_pm"] = data["PM"]
            except Exception:
                pass
            try:
                bdd["ap_pourcent"] = data["Effort"]
            except Exception:
                pass

            if bdd["ap_adresse"] is None or bdd["ap_adresse"] == "":
                del bdd["ap_adresse"]

            if bdd["ap_cable"] is None or bdd["ap_cable"] == "":
                del bdd["ap_cable"]

            if bdd["ap_tvx"] is None or bdd["ap_tvx"] == "":
                del bdd["ap_tvx"]
            else:
                bdd["ap_tvx"] = f"(SELECT tvx_id FROM t_tvx WHERE tvx_choix='{bdd['ap_tvx']}')"

            if bdd["ap_n_app"] is None or bdd["ap_n_app"] == "":
                del bdd["ap_n_app"]
            else:
                bdd["ap_n_app"] = f"(SELECT gs_id FROM t_gespot WHERE gs_nom='{bdd['ap_n_app']}')"

            if bdd["ap_terre"] is None or bdd["ap_terre"] == "":
                del bdd["ap_terre"]
            else:
                bdd["ap_terre"] = f"(SELECT rp_id FROM t_yes_no WHERE rp_nom='{bdd['ap_terre'].upper()}')"

            if bdd["ap_tension"] is None or bdd["ap_tension"] == "":
                del bdd["ap_tension"]
            else:
                bdd["ap_tension"] = f"(SELECT ts_id FROM t_tension WHERE ts_type='{bdd['ap_tension']}')"

            if bdd["ap_obs"] is None or bdd["ap_obs"] == '':
                del bdd["ap_obs"]

            if bdd["ap_hauteur"] is None or bdd["ap_hauteur"] == "":
                del bdd["ap_hauteur"]
            else:
                bdd["ap_hauteur"] = int(bdd["ap_hauteur"])

            if bdd["ap_class"] is None or bdd["ap_class"] == "":
                del bdd["ap_class"]
            else:
                bdd["ap_class"] = f"(SELECT cl_id FROM t_class WHERE cl_nom='{bdd['ap_class']}')"

            if bdd["ap_effortnomi"] is None or bdd["ap_effortnomi"] == "":
                del bdd["ap_effortnomi"]
            else:
                bdd["ap_effortnomi"] = float(bdd["ap_effortnomi"].replace(",", "."))

            if bdd["ap_gespot"] is None or bdd["ap_gespot"] == "":
                del bdd["ap_gespot"]
            else:
                bdd["ap_gespot"] = f"(SELECT gs_id FROM t_gespot WHERE gs_nom='{bdd['ap_gespot']}')"

            if bdd["ap_bv"] is None or bdd["ap_bv"] == "":
                del bdd["ap_bv"]
            else:
                bdd["ap_bv"] = f"(SELECT rp_id FROM t_yes_no WHERE rp_nom='{bdd['ap_bv'].upper()}')"

            if bdd["ap_pourcent"] is None or bdd["ap_pourcent"] == "":
                del bdd["ap_pourcent"]
            else:
                bdd["ap_pourcent"] = float(bdd["ap_pourcent"].replace(",", "."))

            if bdd["ap_artere"] is None or bdd["ap_artere"] == "":
                del bdd["ap_artere"]

            if bdd["ap_pm"] is None or bdd["ap_pm"] == "":
                del bdd["ap_pm"]

            return bdd

        data = req(data=data, date=date)

        for key, item in data.items():
            if isinstance(item, str) and item[0] != "(":
                data[key] = f"'{item}'"
            else:
                data[key] = str(item)
        return data